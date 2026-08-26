"""미국 금융 스트레스 파이프라인의 외부 공식자료 수집·파싱 어댑터.

이 모듈은 자료를 월·주 단위 값으로 정규화하는 일까지만 담당한다.
지수 산식과 Supabase 저장 순서는 ``financial_stress_pipeline``에 둔다.
"""

from __future__ import annotations

import calendar
import csv
import io
import re
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
import requests

from common import fetch_fred_observations


TIMEOUT_SECONDS = 45
OFFICIAL_DATA_HEADERS = {"User-Agent": "MacroWatch/1.0 (+https://hoorash4.github.io/macrowatch/)"}
EBP_CSV_URL = "https://www.federalreserve.gov/econres/notes/feds-notes/ebp_csv.csv"
CMDI_XLSX_URL = "https://www.newyorkfed.org/medialibrary/research/interactives/cmdi/downloads/Market%20CMDI.xlsx"
COURTS_URLS = (
    "https://www.uscourts.gov/sites/default/files/document/bf_f2.1_{period}.xlsx",
    "https://www.uscourts.gov/sites/default/files/data_tables/bf_f2.1_{period}.xlsx",
    "https://www.uscourts.gov/sites/default/files/{publication_year}-{publication_month:02d}/bf_f2.1_{period}.xlsx",
)
MONTH_PATTERN = re.compile(r"Ending\s+([A-Za-z]+)\s+(\d{1,2}),\s+(\d{4})")


def month_start(value: date) -> str:
    return value.replace(day=1).isoformat()


def quarter_ends(start_year: int, start_month: int, end_year: int, end_month: int):
    cursor_year, cursor_month = start_year, ((start_month - 1) // 3 + 1) * 3
    while (cursor_year, cursor_month) <= (end_year, end_month):
        yield cursor_year, cursor_month, calendar.monthrange(cursor_year, cursor_month)[1]
        cursor_month += 3
        if cursor_month > 12:
            cursor_year += 1
            cursor_month = 3


def latest_completed_quarter_end(today: date) -> date:
    quarter_start_month = ((today.month - 1) // 3) * 3 + 1
    if quarter_start_month == 1:
        return date(today.year - 1, 12, 31)
    previous_month = quarter_start_month - 1
    return date(today.year, previous_month, calendar.monthrange(today.year, previous_month)[1])


def _fred_observations(series_id: str, api_key: str, start: date, end: date):
    return fetch_fred_observations(
        series_id,
        api_key,
        start=start.isoformat(),
        end=end.isoformat(),
        timeout=TIMEOUT_SECONDS,
    )


def fetch_fred_monthly(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    values: dict[str, list[float]] = defaultdict(list)
    for observation in _fred_observations(series_id, api_key, start, end):
        raw_value = observation.get("value")
        if raw_value in (None, "."):
            continue
        try:
            values[observation["date"][:7] + "-01"].append(float(raw_value))
        except (KeyError, TypeError, ValueError):
            continue
    return {month: sum(rows) / len(rows) for month, rows in values.items() if rows}


def fetch_fred_month_end(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    """각 달의 마지막 유효 일별 관측값을 반환한다."""
    values: dict[str, tuple[str, float]] = {}
    for observation in _fred_observations(series_id, api_key, start, end):
        raw_value, observed_on = observation.get("value"), observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str):
            continue
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue
        month = observed_on[:7] + "-01"
        if month not in values or observed_on > values[month][0]:
            values[month] = (observed_on, value)
    return {month: value for month, (_observed_on, value) in values.items()}


def fetch_fred_latest(series_id: str, api_key: str, start: date, end: date) -> tuple[date, float] | None:
    latest: tuple[date, float] | None = None
    for observation in _fred_observations(series_id, api_key, start, end):
        raw_value, observed_on = observation.get("value"), observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str):
            continue
        try:
            candidate = (date.fromisoformat(observed_on), float(raw_value))
        except (TypeError, ValueError):
            continue
        if latest is None or candidate[0] > latest[0]:
            latest = candidate
    return latest


def fetch_fred_week_end(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    """금요일 종료 주간별 마지막 유효 관측값을 반환한다."""
    values: dict[str, tuple[str, float]] = {}
    for observation in _fred_observations(series_id, api_key, start, end):
        raw_value, observed_on = observation.get("value"), observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str):
            continue
        try:
            observed_date, value = date.fromisoformat(observed_on), float(raw_value)
        except (TypeError, ValueError):
            continue
        week = (observed_date + timedelta(days=4 - observed_date.weekday())).isoformat()
        if week not in values or observed_on > values[week][0]:
            values[week] = (observed_on, value)
    return {week: value for week, (_observed_on, value) in values.items()}


def fetch_ebp_monthly(start: date, end: date) -> dict[str, float]:
    """연준의 월별 Excess Bond Premium CSV를 읽는다."""
    response = requests.get(EBP_CSV_URL, headers=OFFICIAL_DATA_HEADERS, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()
    reader = csv.DictReader(io.StringIO(response.content.decode("utf-8-sig", errors="replace")))
    values: dict[str, float] = {}
    for row in reader:
        normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
        raw_date = normalized.get("date") or normalized.get("observation_date")
        if raw_date is None:
            raw_date = next((value for key, value in normalized.items() if "date" in key), None)
        raw_value = normalized.get("ebp")
        if raw_value is None:
            raw_value = next(
                (value for key, value in normalized.items() if key.endswith("ebp") or "excess bond premium" in key),
                None,
            )
        observed_on = _parse_date(raw_date)
        try:
            value = float(str(raw_value))
        except (TypeError, ValueError):
            continue
        if observed_on is not None and start <= observed_on <= end:
            values[month_start(observed_on)] = value
    if not values:
        raise RuntimeError("연준 EBP CSV에서 월별 값을 찾지 못했습니다.")
    return values


def _parse_date(raw_value: object) -> date | None:
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    raw_text = str(raw_value).split(" ")[0]
    for pattern in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_text, pattern).date()
        except ValueError:
            continue
    return None


def fetch_cmdi_monthly(start: date, end: date) -> dict[str, float]:
    """뉴욕연은 Market CMDI 워크북을 메모리에서 읽는다."""
    response = requests.get(CMDI_XLSX_URL, headers=OFFICIAL_DATA_HEADERS, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    values: dict[str, tuple[date, float]] = {}
    try:
        for sheet in workbook.worksheets:
            header_row = date_column = value_column = None
            for row_number, header in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=True), start=1):
                labels = [str(value or "").strip().lower() for value in header]
                try:
                    date_column = next(
                        index for index, label in enumerate(labels)
                        if label == "date" or "date" in label or label in {"eow_friday", "week_end", "week ending"}
                    )
                    value_column = next(
                        index for index, label in enumerate(labels)
                        if ("market" in label and "cmdi" in label) or label == "market"
                    )
                    header_row = row_number
                    break
                except StopIteration:
                    continue
            if header_row is None or date_column is None or value_column is None:
                continue
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if len(row) <= max(date_column, value_column):
                    continue
                observed_on = _parse_date(row[date_column])
                try:
                    value = float(row[value_column])
                except (TypeError, ValueError):
                    continue
                if observed_on is not None and start <= observed_on <= end:
                    month = month_start(observed_on)
                    if month not in values or observed_on > values[month][0]:
                        values[month] = (observed_on, value)
    finally:
        workbook.close()
    if not values:
        raise RuntimeError("뉴욕연은 CMDI XLSX에서 Market CMDI 값을 찾지 못했습니다.")
    return {month: value for month, (_observed_on, value) in values.items()}


def fetch_court_workbook(year: int, month: int, day: int) -> bytes:
    period = f"{month:02d}{day:02d}.{year}"
    publication_year, publication_month = year, month + 1
    if publication_month == 13:
        publication_year, publication_month = year + 1, 1
    for template in COURTS_URLS:
        response = requests.get(
            template.format(
                period=period,
                publication_year=publication_year,
                publication_month=publication_month,
            ),
            timeout=TIMEOUT_SECONDS,
        )
        if response.ok:
            return response.content
    raise RuntimeError(f"법원 F-2 월간 XLSX를 찾지 못했습니다: {year}-{month:02d}")


def parse_business_filings(workbook_bytes: bytes) -> dict[str, int]:
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    filings: dict[str, int] = {}
    try:
        for sheet in workbook.worksheets:
            if "(" in sheet.title:
                continue
            match = MONTH_PATTERN.search(str(sheet.cell(2, 1).value or ""))
            if not match:
                continue
            total_row = next(
                (row for row in range(1, sheet.max_row + 1) if str(sheet.cell(row, 1).value or "").strip() == "Total"),
                None,
            )
            if total_row is None:
                raise RuntimeError(f"법원 F-2 표에서 Total 행을 찾지 못했습니다: {sheet.title}")
            try:
                filing_count = int(str(sheet.cell(total_row, 7).value).replace(",", ""))
            except (TypeError, ValueError) as error:
                raise RuntimeError(f"법원 F-2 사업체 파산보호 건수가 올바르지 않습니다: {sheet.title}") from error
            month_number = list(calendar.month_name).index(match.group(1))
            filings[f"{match.group(3)}-{month_number:02d}-01"] = filing_count
    finally:
        workbook.close()
    if not filings:
        raise RuntimeError("법원 F-2 XLSX에서 월별 사업체 파산보호 신청 건수를 찾지 못했습니다.")
    return filings


def collect_business_filings(start: date, end: date) -> dict[str, int]:
    filings: dict[str, int] = {}
    for year, month, day in quarter_ends(start.year, start.month, end.year, end.month):
        try:
            workbook = fetch_court_workbook(year, month, day)
        except RuntimeError as error:
            print(f"skipped_court_report={year}-{month:02d} reason={error}")
            continue
        for month_key, value in parse_business_filings(workbook).items():
            if start.isoformat()[:7] <= month_key[:7] <= end.isoformat()[:7]:
                filings[month_key] = value
    return filings
