"""Collect monthly U.S. credit-stress indicators without retaining source files."""

from __future__ import annotations

import argparse
import calendar
import csv
import io
import os
import re
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
import requests


FRED_URL = "https://api.stlouisfed.org/fred/series/observations"
COURTS_URLS = (
    "https://www.uscourts.gov/sites/default/files/document/bf_f2.1_{period}.xlsx",
    "https://www.uscourts.gov/sites/default/files/data_tables/bf_f2.1_{period}.xlsx",
    "https://www.uscourts.gov/sites/default/files/{publication_year}-{publication_month:02d}/bf_f2.1_{period}.xlsx",
)
HIGH_YIELD_SERIES = "BAMLH0A0HYM2"
FINANCIAL_CONDITIONS_SERIES = "NFCICREDIT"
FINANCIAL_RISK_SERIES = "NFCIRISK"
NONFINANCIAL_LEVERAGE_SERIES = "NFCINONFINLEVERAGE"
SLOOS_SERIES = "DRTSCILM"
SP500_SERIES = "SP500"
COMMERCIAL_PAPER_SERIES = "DCPN3M"
THREE_MONTH_TREASURY_SERIES = "DGS3MO"
MONTH_PATTERN = re.compile(r"Ending\s+([A-Za-z]+)\s+(\d{1,2}),\s+(\d{4})")
TIMEOUT_SECONDS = 45
INDEX_HISTORY_YEARS = 3
OFFICIAL_DATA_HEADERS = {"User-Agent": "MacroWatch/1.0 (+https://hoorash4.github.io/macrowatch/)"}
EBP_CSV_URL = "https://www.federalreserve.gov/econres/notes/feds-notes/ebp_csv.csv"
CMDI_XLSX_URL = "https://www.newyorkfed.org/medialibrary/research/interactives/cmdi/downloads/Market%20CMDI.xlsx"
STRESS_COMPONENTS = (
    "financial_conditions_risk_index",
    "high_yield_oas_pct",
    "financial_conditions_credit_index",
    "corporate_bond_market_distress_index",
    "sloos_tightening_pct",
    "business_bankruptcy_filings",
)
LEAD_COMPONENT_WEIGHTS = {
    "high_yield": 0.20,
    "financial_conditions": 0.20,
    "short_term_funding_spread": 0.20,
    "nonfinancial_leverage": 0.20,
    "excess_bond_premium": 0.20,
}
SHORT_TERM_FUNDING_FLOOR = 0.10
SHORT_TERM_FUNDING_REFERENCE = 0.60
SHORT_TERM_FUNDING_CHANGE_REFERENCE = 0.30
# Fixed 0-to-100 reference ranges. These never roll with incoming data; values
# above the reference range deliberately remain above 100 to preserve stress
# severity during future extremes.
FIXED_COMPONENT_SCALES = {
    "high_yield_oas_pct": (2.0, 20.0),
    "financial_conditions_credit_index": (-0.5, 2.0),
    "financial_conditions_risk_index": (-0.5, 2.0),
    "corporate_bond_market_distress_index": (0.0, 1.0),
    "excess_bond_premium": (-1.0, 4.0),
    "nonfinancial_leverage_index": (-1.5, 2.0),
    "business_bankruptcy_filings": (1000.0, 5000.0),
    "sloos_tightening_pct": (0.0, 100.0),
}


def month_start(value: date) -> str:
    return value.replace(day=1).isoformat()


def quarter_ends(start_year: int, start_month: int, end_year: int, end_month: int):
    cursor_year, cursor_month = start_year, ((start_month - 1) // 3 + 1) * 3
    while (cursor_year, cursor_month) <= (end_year, end_month):
        yield cursor_year, cursor_month, calendar.monthrange(cursor_year, cursor_month)[1]
        cursor_month += 3
        if cursor_month > 12:
            cursor_year += 1
            cursor_month = 3


def latest_completed_quarter_end(today: date) -> date:
    quarter_start_month = ((today.month - 1) // 3) * 3 + 1
    if quarter_start_month == 1:
        return date(today.year - 1, 12, 31)
    previous_month = quarter_start_month - 1
    return date(today.year, previous_month, calendar.monthrange(today.year, previous_month)[1])


def fetch_fred_monthly(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    response = requests.get(
        FRED_URL,
        params={
            "series_id": series_id,
            "api_key": api_key,
            "file_type": "json",
            "observation_start": start.isoformat(),
            "observation_end": end.isoformat(),
            "limit": "100000",
        },
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    values: dict[str, list[float]] = defaultdict(list)
    for observation in response.json().get("observations", []):
        raw_value = observation.get("value")
        if raw_value in (None, "."):
            continue
        try:
            values[observation["date"][:7] + "-01"].append(float(raw_value))
        except (KeyError, TypeError, ValueError):
            continue
    return {month: sum(rows) / len(rows) for month, rows in values.items() if rows}


def fetch_fred_quarterly_to_months(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    """Expand a quarterly observation into the three months it measures.

    SLOOS observations are dated on the first day after their reference
    quarter.  A value dated 2026-04-01 therefore belongs to Jan--Mar 2026,
    not to the publication month.  The monthly index is revised from
    provisional to confirmed when that observation becomes available.
    """
    response = requests.get(
        FRED_URL,
        params={
            "series_id": series_id,
            "api_key": api_key,
            "file_type": "json",
            "observation_start": (start - timedelta(days=100)).isoformat(),
            "observation_end": (end + timedelta(days=100)).isoformat(),
            "limit": "100000",
        },
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    values: dict[str, float] = {}
    for observation in response.json().get("observations", []):
        raw_value, observed_on = observation.get("value"), observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str):
            continue
        try:
            value = float(raw_value)
            quarter_end = date.fromisoformat(observed_on) - timedelta(days=1)
        except (TypeError, ValueError):
            continue
        for offset in range(3):
            year = quarter_end.year
            month_number = quarter_end.month - offset
            while month_number <= 0:
                year -= 1
                month_number += 12
            month = date(year, month_number, 1)
            if start <= month <= end:
                values[month.isoformat()] = value
    return values


def fetch_fred_month_end(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    """Return the final available daily observation for each calendar month."""
    response = requests.get(
        FRED_URL,
        params={
            "series_id": series_id,
            "api_key": api_key,
            "file_type": "json",
            "observation_start": start.isoformat(),
            "observation_end": end.isoformat(),
            "limit": "100000",
        },
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    month_end: dict[str, tuple[str, float]] = {}
    for observation in response.json().get("observations", []):
        raw_value = observation.get("value")
        observed_on = observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str):
            continue
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue
        month = observed_on[:7] + "-01"
        if month not in month_end or observed_on > month_end[month][0]:
            month_end[month] = (observed_on, value)
    return {month: value for month, (_observed_on, value) in month_end.items()}


def fetch_fred_latest(series_id: str, api_key: str, start: date, end: date) -> tuple[date, float] | None:
    response = requests.get(
        FRED_URL,
        params={"series_id": series_id, "api_key": api_key, "file_type": "json", "observation_start": start.isoformat(), "observation_end": end.isoformat(), "limit": "100000"},
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    latest: tuple[date, float] | None = None
    for observation in response.json().get("observations", []):
        raw_value, observed_on = observation.get("value"), observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str):
            continue
        try:
            candidate = (date.fromisoformat(observed_on), float(raw_value))
        except (TypeError, ValueError):
            continue
        if latest is None or candidate[0] > latest[0]:
            latest = candidate
    return latest


def fetch_fred_week_end(series_id: str, api_key: str, start: date, end: date) -> dict[str, float]:
    """Return the final available observation for each Friday-ended week."""
    response = requests.get(FRED_URL, params={"series_id": series_id, "api_key": api_key, "file_type": "json", "observation_start": start.isoformat(), "observation_end": end.isoformat(), "limit": "100000"}, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()
    values: dict[str, tuple[str, float]] = {}
    for observation in response.json().get("observations", []):
        raw_value, observed_on = observation.get("value"), observation.get("date")
        if raw_value in (None, ".") or not isinstance(observed_on, str): continue
        try: observed_date, value = date.fromisoformat(observed_on), float(raw_value)
        except (TypeError, ValueError): continue
        week = (observed_date + timedelta(days=4 - observed_date.weekday())).isoformat()
        if week not in values or observed_on > values[week][0]: values[week] = (observed_on, value)
    return {week: value for week, (_observed_on, value) in values.items()}


def fetch_ebp_monthly(start: date, end: date) -> dict[str, float]:
    """Read the Federal Reserve Board's monthly Excess Bond Premium CSV."""
    response = requests.get(EBP_CSV_URL, headers=OFFICIAL_DATA_HEADERS, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()
    reader = csv.DictReader(io.StringIO(response.content.decode("utf-8-sig", errors="replace")))
    values: dict[str, float] = {}
    for row in reader:
        normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
        raw_date = normalized.get("date") or normalized.get("observation_date")
        if raw_date is None:
            raw_date = next((value for key, value in normalized.items() if "date" in key), None)
        raw_value = normalized.get("ebp")
        if raw_value is None:
            raw_value = next((value for key, value in normalized.items() if key.endswith("ebp") or "excess bond premium" in key), None)
        raw_date_text = str(raw_date).split(" ")[0]
        observed_on = None
        for pattern in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                observed_on = datetime.strptime(raw_date_text, pattern).date()
                break
            except ValueError:
                continue
        try:
            if observed_on is None:
                continue
            value = float(str(raw_value))
        except (TypeError, ValueError):
            continue
        if start <= observed_on <= end:
            values[month_start(observed_on)] = value
    if not values:
        raise RuntimeError("연준 EBP CSV에서 월별 값을 찾지 못했습니다.")
    return values


def fetch_cmdi_monthly(start: date, end: date) -> dict[str, float]:
    """Read the New York Fed Market CMDI workbook without retaining the file."""
    response = requests.get(CMDI_XLSX_URL, headers=OFFICIAL_DATA_HEADERS, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    values: dict[str, tuple[date, float]] = {}
    for sheet in workbook.worksheets:
        header_row = date_column = value_column = None
        for row_number, header in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=True), start=1):
            labels = [str(value or "").strip().lower() for value in header]
            try:
                date_column = next(
                    index
                    for index, label in enumerate(labels)
                    if label == "date" or "date" in label or label in {"eow_friday", "week_end", "week ending"}
                )
                value_column = next(index for index, label in enumerate(labels) if ("market" in label and "cmdi" in label) or label == "market")
                header_row = row_number
                break
            except StopIteration:
                continue
        if header_row is None or date_column is None or value_column is None:
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if len(row) <= max(date_column, value_column):
                continue
            raw_date, raw_value = row[date_column], row[value_column]
            if isinstance(raw_date, datetime):
                observed_on = raw_date.date()
            elif isinstance(raw_date, date):
                observed_on = raw_date
            else:
                raw_date_text = str(raw_date).split(" ")[0]
                observed_on = None
                for pattern in ("%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        observed_on = datetime.strptime(raw_date_text, pattern).date()
                        break
                    except ValueError:
                        continue
                if observed_on is None:
                    continue
            try:
                value = float(raw_value)
            except (TypeError, ValueError):
                continue
            if start <= observed_on <= end:
                month = month_start(observed_on)
                if month not in values or observed_on > values[month][0]:
                    values[month] = (observed_on, value)
    if not values:
        raise RuntimeError("뉴욕연은 CMDI XLSX에서 Market CMDI 값을 찾지 못했습니다.")
    return {month: value for month, (_observed_on, value) in values.items()}


def monthly_values_for_weeks(values: dict[str, float], weeks: list[str]) -> dict[str, float]:
    """Apply a released monthly value consistently to every weekly bucket in that month."""
    return {week: values[week[:7] + "-01"] for week in weeks if week[:7] + "-01" in values}


def carry_forward_values(values: dict[str, float], periods: list[str]) -> dict[str, float]:
    """Fill an unreleased period with the latest actual observation."""
    carried: dict[str, float] = {}
    previous: float | None = None
    for period in sorted(periods):
        current = values.get(period)
        if current is not None:
            previous = current
        if previous is not None:
            carried[period] = previous
    return carried


def build_weekly_lead(
    high_yield: dict[str, float],
    conditions: dict[str, float],
    funding: dict[str, float],
    leverage: dict[str, float],
    excess_bond_premium: dict[str, float],
) -> list[dict[str, object]]:
    weeks = sorted(set(high_yield) | set(conditions) | set(funding) | set(leverage))
    excess_bond_premium = monthly_values_for_weeks(excess_bond_premium, weeks)
    raw_sources = (high_yield, conditions, funding, leverage, excess_bond_premium)
    high_yield, conditions, funding, leverage, excess_bond_premium = (
        carry_forward_values(values, weeks)
        for values in raw_sources
    )
    rows, previous = [], None
    for week in weeks:
        hy, condition, spread, leverage_value, ebp = (
            high_yield.get(week),
            conditions.get(week),
            funding.get(week),
            leverage.get(week),
            excess_bond_premium.get(week),
        )
        if not all(isinstance(value, (int, float)) for value in (hy, condition, spread, leverage_value, ebp)):
            continue
        level = (
            fixed_stress_score(float(hy), "high_yield_oas_pct") * LEAD_COMPONENT_WEIGHTS["high_yield"]
            + fixed_stress_score(float(condition), "financial_conditions_credit_index") * LEAD_COMPONENT_WEIGHTS["financial_conditions"]
            + positive_score(float(spread) - SHORT_TERM_FUNDING_FLOOR, SHORT_TERM_FUNDING_REFERENCE) * LEAD_COMPONENT_WEIGHTS["short_term_funding_spread"]
            + fixed_stress_score(float(leverage_value), "nonfinancial_leverage_index") * LEAD_COMPONENT_WEIGHTS["nonfinancial_leverage"]
            + fixed_stress_score(float(ebp), "excess_bond_premium") * LEAD_COMPONENT_WEIGHTS["excess_bond_premium"]
        )
        momentum = None if previous is None else (
            signed_score(float(hy) - previous[0], 1.0) * LEAD_COMPONENT_WEIGHTS["high_yield"]
            + signed_score(float(condition) - previous[1], 0.5) * LEAD_COMPONENT_WEIGHTS["financial_conditions"]
            + signed_score(float(spread) - previous[2], SHORT_TERM_FUNDING_CHANGE_REFERENCE) * LEAD_COMPONENT_WEIGHTS["short_term_funding_spread"]
            + signed_score(float(leverage_value) - previous[3], 0.5) * LEAD_COMPONENT_WEIGHTS["nonfinancial_leverage"]
            + signed_score(float(ebp) - previous[4], 0.5) * LEAD_COMPONENT_WEIGHTS["excess_bond_premium"]
        )
        rows.append({
            "week": week,
            "lead_index": round(level, 2),
            "lead_momentum": None if momentum is None else round(momentum, 2),
            "leverage_signal": round(fixed_stress_score(float(leverage_value), "nonfinancial_leverage_index"), 2),
            "is_provisional": any(week not in source for source in raw_sources),
        })
        previous = (float(hy), float(condition), float(spread), float(leverage_value), float(ebp))
    return rows


def fetch_court_workbook(year: int, month: int, day: int) -> bytes:
    period = f"{month:02d}{day:02d}.{year}"
    publication_year, publication_month = year, month + 1
    if publication_month == 13:
        publication_year, publication_month = year + 1, 1
    for template in COURTS_URLS:
        response = requests.get(
            template.format(
                period=period,
                publication_year=publication_year,
                publication_month=publication_month,
            ),
            timeout=TIMEOUT_SECONDS,
        )
        if response.ok:
            return response.content
    raise RuntimeError(f"법원 F-2 월간 XLSX를 찾지 못했습니다: {year}-{month:02d}")


def parse_business_filings(workbook_bytes: bytes) -> dict[str, int]:
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    filings: dict[str, int] = {}
    for sheet in workbook.worksheets:
        # F-2 also contains chapter-detail sheets named "(9, 12, 15)".
        # The monthly total is only present in the base F-2 sheet.
        if "(" in sheet.title:
            continue
        title = str(sheet.cell(2, 1).value or "")
        match = MONTH_PATTERN.search(title)
        if not match:
            continue
        total_row = next(
            (row for row in range(1, sheet.max_row + 1) if str(sheet.cell(row, 1).value or "").strip() == "Total"),
            None,
        )
        if total_row is None:
            raise RuntimeError(f"법원 F-2 표에서 Total 행을 찾지 못했습니다: {sheet.title}")
        value = sheet.cell(total_row, 7).value  # Predominant Nature of Debt: Business, All Chapters
        try:
            filing_count = int(str(value).replace(",", ""))
        except (TypeError, ValueError) as error:
            raise RuntimeError(f"법원 F-2 사업체 파산보호 건수가 올바르지 않습니다: {sheet.title}") from error
        month_number = list(calendar.month_name).index(match.group(1))
        filings[f"{match.group(3)}-{month_number:02d}-01"] = filing_count
    if not filings:
        raise RuntimeError("법원 F-2 XLSX에서 월별 사업체 파산보호 신청 건수를 찾지 못했습니다.")
    return filings


def collect_business_filings(start: date, end: date) -> dict[str, int]:
    filings: dict[str, int] = {}
    for year, month, day in quarter_ends(start.year, start.month, end.year, end.month):
        try:
            workbook = fetch_court_workbook(year, month, day)
        except RuntimeError as error:
            # The court publishes each three-month report after its period ends.
            # Keep the available history intact while a newly expected report is pending.
            print(f"skipped_court_report={year}-{month:02d} reason={error}")
            continue
        for month_key, value in parse_business_filings(workbook).items():
            if start.isoformat()[:7] <= month_key[:7] <= end.isoformat()[:7]:
                filings[month_key] = value
    return filings


def upsert_rows(rows: list[dict[str, object]], supabase_url: str, service_role_key: str) -> None:
    response = requests.post(
        f"{supabase_url.rstrip('/')}/rest/v1/us_credit_stress_monthly?on_conflict=month",
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        json=rows,
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()


def fixed_stress_score(value: float, key: str) -> float:
    floor, reference = FIXED_COMPONENT_SCALES[key]
    return max(0.0, ((value - floor) / (reference - floor)) * 100.0)


def smoothed_filings(rows: list[dict[str, object]]) -> tuple[dict[str, float], set[str]]:
    """Return a trailing filing average and the months backed by published reports."""
    observed: list[float] = []
    values: dict[str, float] = {}
    confirmed: set[str] = set()
    for row in sorted(rows, key=lambda item: str(item["month"])):
        raw_value = row.get("business_bankruptcy_filings")
        if isinstance(raw_value, (int, float)) and raw_value > 0:
            observed.append(float(raw_value))
            values[str(row["month"])] = sum(observed[-3:]) / min(3, len(observed))
            confirmed.add(str(row["month"]))
        elif observed:
            # A court report is not yet available. Preserve the latest published
            # trend instead of treating an unreported month as zero or changing
            # component weights.
            values[str(row["month"])] = sum(observed[-3:]) / min(3, len(observed))
    return values, confirmed


def positive_score(value: float, reference: float) -> float:
    return max(0.0, value / reference * 100.0)


def signed_score(value: float, reference: float) -> float:
    return value / reference * 100.0


def build_market_stress_lead(
    rows: list[dict[str, object]],
    short_term_funding_spread: dict[str, float],
) -> tuple[dict[str, float], dict[str, float]]:
    """Return independent level and month-over-month market-stress signals."""
    ordered_rows = sorted(rows, key=lambda row: str(row["month"]))
    lead_scores: dict[str, float] = {}
    momentum_scores: dict[str, float] = {}
    for index, row in enumerate(ordered_rows):
        month = str(row["month"])
        try:
            funding_spread = float(short_term_funding_spread[month])
            component_scores = {
                "high_yield": fixed_stress_score(float(row["high_yield_oas_pct"]), "high_yield_oas_pct"),
                "financial_conditions": fixed_stress_score(float(row["financial_conditions_credit_index"]), "financial_conditions_credit_index"),
                "short_term_funding_spread": positive_score(
                    funding_spread - SHORT_TERM_FUNDING_FLOOR,
                    SHORT_TERM_FUNDING_REFERENCE,
                ),
            }
        except (KeyError, TypeError, ValueError):
            continue
        available_weights = {
            key: LEAD_COMPONENT_WEIGHTS[key]
            for key in component_scores
        }
        available_weight_total = sum(available_weights.values())
        lead_scores[month] = round(
            sum(component_scores[key] * weight for key, weight in available_weights.items())
            / available_weight_total,
            2,
        )
        if index == 0:
            continue
        previous = ordered_rows[index - 1]
        try:
            momentum_components = {
                "high_yield": signed_score(
                    float(row["high_yield_oas_pct"]) - float(previous["high_yield_oas_pct"]),
                    1.0,
                ),
                "financial_conditions": signed_score(
                    float(row["financial_conditions_credit_index"]) - float(previous["financial_conditions_credit_index"]),
                    0.5,
                ),
                "short_term_funding_spread": signed_score(
                    funding_spread - float(short_term_funding_spread[str(previous["month"])]),
                    SHORT_TERM_FUNDING_CHANGE_REFERENCE,
                ),
            }
        except (KeyError, TypeError, ValueError):
            continue
        momentum_scores[month] = round(
            sum(momentum_components[key] * weight for key, weight in available_weights.items())
            / available_weight_total,
            2,
        )
    return lead_scores, momentum_scores


def build_market_stress_index(
    rows: list[dict[str, object]],
    today: date,
    sp500_month_end: dict[str, float],
    short_term_funding_spread: dict[str, float],
) -> list[dict[str, object]]:
    index_start = date(today.year - INDEX_HISTORY_YEARS, today.month, 1).isoformat()
    recent_rows = [row for row in rows if str(row["month"]) >= index_start]
    filings, confirmed_filings = smoothed_filings(recent_rows)
    lead_scores, momentum_scores = build_market_stress_lead(recent_rows, short_term_funding_spread)
    months = [str(row["month"]) for row in recent_rows]
    raw_component_values: dict[str, dict[str, float]] = {
        "business_bankruptcy_filings": filings,
    }
    for key in STRESS_COMPONENTS:
        if key == "business_bankruptcy_filings":
            continue
        raw_component_values[key] = {
            str(row["month"]): float(row[key])
            for row in recent_rows
            if isinstance(row.get(key), (int, float))
        }
    component_values = {
        key: carry_forward_values(values, months)
        for key, values in raw_component_values.items()
    }
    component_scores = {
        key: {month: fixed_stress_score(value, key) for month, value in values.items()}
        for key, values in component_values.items()
        if values
    }
    index_rows: list[dict[str, object]] = []
    for row in recent_rows:
        month = str(row["month"])
        if any(month not in component_scores.get(key, {}) for key in STRESS_COMPONENTS):
            continue
        score = sum(component_scores[key][month] for key in STRESS_COMPONENTS) / len(STRESS_COMPONENTS)
        index_rows.append(
            {
                "month": month,
                "stress_index": round(score, 2),
                "is_provisional": (
                    month not in confirmed_filings
                    or any(month not in raw_component_values.get(key, {}) for key in STRESS_COMPONENTS)
                ),
                "sp500_month_end_close": sp500_month_end.get(month),
                "lead_index": lead_scores.get(month),
                "lead_momentum": momentum_scores.get(month),
            }
        )
    return index_rows


def upsert_market_stress_index(rows: list[dict[str, object]], supabase_url: str, service_role_key: str) -> None:
    if not rows:
        return
    response = requests.post(
        f"{supabase_url.rstrip('/')}/rest/v1/us_market_stress_index_monthly?on_conflict=month",
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        json=rows,
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()


def upsert_weekly_lead(rows: list[dict[str, object]], supabase_url: str, service_role_key: str) -> None:
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/us_market_stress_lead_weekly?on_conflict=week"
    headers = {"apikey": service_role_key, "Authorization": f"Bearer {service_role_key}", "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates,return=minimal"}
    response = requests.post(endpoint, headers=headers, json=rows, timeout=TIMEOUT_SECONDS)
    if response.status_code == 400 and ("leverage_signal" in response.text or "is_provisional" in response.text):
        # Keep the established lead series updating while the additive database
        # migration is still being applied by the production integration.
        legacy_rows = [{key: value for key, value in row.items() if key not in {"leverage_signal", "is_provisional"}} for row in rows]
        response = requests.post(endpoint, headers=headers, json=legacy_rows, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()


def upsert_latest_credit_stress(row: dict[str, object], supabase_url: str, service_role_key: str) -> None:
    response = requests.post(
        f"{supabase_url.rstrip('/')}/rest/v1/us_credit_stress_latest?on_conflict=singleton",
        headers={"apikey": service_role_key, "Authorization": f"Bearer {service_role_key}", "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates,return=minimal"},
        json=row,
        timeout=TIMEOUT_SECONDS,
    )
    response.raise_for_status()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--years", type=int, default=3)
    args = parser.parse_args()
    if args.years < 1 or args.years > 10:
        raise SystemExit("--years 값은 1~10 사이여야 합니다.")

    fred_api_key = os.environ["FRED_API_KEY"]
    supabase_url = os.environ["SUPABASE_URL"]
    service_role_key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    today = date.today()
    start = date(today.year - args.years, today.month, 1)
    end = date(today.year, today.month, 1)

    high_yield = fetch_fred_monthly(HIGH_YIELD_SERIES, fred_api_key, start, end)
    financial_conditions = fetch_fred_monthly(FINANCIAL_CONDITIONS_SERIES, fred_api_key, start, end)
    financial_risk = fetch_fred_monthly(FINANCIAL_RISK_SERIES, fred_api_key, start, end)
    sloos = fetch_fred_quarterly_to_months(SLOOS_SERIES, fred_api_key, start, end)
    excess_bond_premium = fetch_ebp_monthly(start, end)
    cmdi = fetch_cmdi_monthly(start, end)
    sp500_month_end = fetch_fred_month_end(SP500_SERIES, fred_api_key, start, end)
    commercial_paper = fetch_fred_monthly(COMMERCIAL_PAPER_SERIES, fred_api_key, start, end)
    three_month_treasury = fetch_fred_monthly(THREE_MONTH_TREASURY_SERIES, fred_api_key, start, end)
    short_term_funding_spread = {
        month: commercial_paper[month] - three_month_treasury[month]
        for month in commercial_paper.keys() & three_month_treasury.keys()
    }
    business_filings = collect_business_filings(start, latest_completed_quarter_end(today))
    months = sorted(
        set(high_yield) | set(financial_conditions) | set(financial_risk)
        | set(sloos) | set(excess_bond_premium) | set(cmdi) | set(business_filings)
    )
    short_term_funding_spread = carry_forward_values(short_term_funding_spread, months)
    rows = [
        {
            "month": month,
            "high_yield_oas_pct": high_yield.get(month),
            "financial_conditions_credit_index": financial_conditions.get(month),
            "financial_conditions_risk_index": financial_risk.get(month),
            "excess_bond_premium": excess_bond_premium.get(month),
            "corporate_bond_market_distress_index": cmdi.get(month),
            "sloos_tightening_pct": sloos.get(month),
            "business_bankruptcy_filings": business_filings.get(month),
        }
        for month in months
    ]
    if not rows:
        raise RuntimeError("저장할 월별 신용 스트레스 데이터가 없습니다.")
    latest_start = today - timedelta(days=60)
    latest_high_yield = fetch_fred_latest(HIGH_YIELD_SERIES, fred_api_key, latest_start, today)
    latest_conditions = fetch_fred_latest(FINANCIAL_CONDITIONS_SERIES, fred_api_key, latest_start, today)
    latest_risk = fetch_fred_latest(FINANCIAL_RISK_SERIES, fred_api_key, latest_start, today)
    latest_dates = [source[0] for source in (latest_high_yield, latest_conditions, latest_risk) if source is not None]
    latest_month = month_start(max(latest_dates)) if latest_dates else None
    if latest_month and latest_high_yield and latest_conditions and latest_risk:
        latest_values = {
            "high_yield_oas_pct": latest_high_yield[1],
            "financial_conditions_credit_index": latest_conditions[1],
            "financial_conditions_risk_index": latest_risk[1],
            "excess_bond_premium": excess_bond_premium.get(latest_month),
            "corporate_bond_market_distress_index": cmdi.get(latest_month),
        }
        matching_row = next((row for row in rows if row["month"] == latest_month), None)
        if matching_row is None:
            rows.append({
                "month": latest_month,
                **latest_values,
                "sloos_tightening_pct": None,
                "business_bankruptcy_filings": None,
            })
        else:
            matching_row.update(latest_values)
    index_rows_input = rows
    upsert_rows(rows, supabase_url, service_role_key)
    index_rows = build_market_stress_index(
        index_rows_input,
        today,
        sp500_month_end,
        short_term_funding_spread,
    )
    upsert_market_stress_index(index_rows, supabase_url, service_role_key)
    weekly_high_yield = fetch_fred_week_end(HIGH_YIELD_SERIES, fred_api_key, start, today)
    weekly_conditions = fetch_fred_week_end(FINANCIAL_CONDITIONS_SERIES, fred_api_key, start, today)
    weekly_leverage = fetch_fred_week_end(NONFINANCIAL_LEVERAGE_SERIES, fred_api_key, start, today)
    weekly_cp = fetch_fred_week_end(COMMERCIAL_PAPER_SERIES, fred_api_key, start, today)
    weekly_treasury = fetch_fred_week_end(THREE_MONTH_TREASURY_SERIES, fred_api_key, start, today)
    weekly_funding = {week: weekly_cp[week] - weekly_treasury[week] for week in weekly_cp.keys() & weekly_treasury.keys()}
    weekly_rows = build_weekly_lead(
        weekly_high_yield,
        weekly_conditions,
        weekly_funding,
        weekly_leverage,
        excess_bond_premium,
    )
    upsert_weekly_lead(weekly_rows, supabase_url, service_role_key)
    if latest_dates:
        upsert_latest_credit_stress({
            "singleton": True,
            "as_of": max(latest_dates).isoformat(),
            "high_yield_oas_pct": latest_high_yield[1] if latest_high_yield else None,
            "financial_conditions_credit_index": latest_conditions[1] if latest_conditions else None,
        }, supabase_url, service_role_key)
    print(
        f"upserted_months={len(rows)} market_stress_index={len(index_rows)} "
        f"business_filings={len(business_filings)} high_yield={len(high_yield)} "
        f"nfci_credit={len(financial_conditions)} nfci_risk={len(financial_risk)} "
        f"ebp_months={len(excess_bond_premium)} cmdi_months={len(cmdi)} "
        f"sloos_months={len(sloos)} sp500={len(sp500_month_end)} "
        f"short_funding_spread={len(short_term_funding_spread)} "
        f"weekly_leverage={len(weekly_leverage)}"
    )


if __name__ == "__main__":
    main()
